# %% [markdown]
# # Statistical Arbitrage for Sports Betting

# %% [markdown]
# Using: Live Sports Odds API
# Documentation Link: https://the-odds-api.com/ 

# %% [markdown]
# This program will look for statistical arbitrage opportunities in the upcoming eight games across all sports.

# %% [markdown]
# ### Importing Dependencies and Acquiring API Key

# %%
import requests
# import xlsxwriter
import pandas as pd
import numpy as np
# import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, numbers
from dotenv import load_dotenv
import os
load_dotenv('.env')
def get_upcoming_odds():
    API_KEY = os.getenv('API_KEY')
    SPORT = 'upcoming'
    REGIONS = 'us'
    MARKETS = 'h2h'
    ODDS_FORMAT = 'decimal'
    DATE_FORMAT = 'iso'
    odds_response = requests.get(
        f'https://api.the-odds-api.com/v4/sports/{SPORT}/odds',
        params={
            'api_key': API_KEY,
            'regions': REGIONS,
            'markets': MARKETS,
            'oddsFormat': ODDS_FORMAT,
            'dateFormat': DATE_FORMAT,
        }
    ).json()
    return odds_response

BOOKMAKER_INDEX = 0
NAME_INDEX = 1
ODDS_INDEX = 2
FIRST = 0

class Event:
    def __init__(self, data):
        self.data = data
        self.sport_key = data['sport_key']
        self.id = data['id']
       
    def find_best_odds(self):
        best_odds = []
        bookmakers = self.data.get('bookmakers', [])
        
        for bookmaker in bookmakers:
            markets = bookmaker.get('markets', [])
            if markets:
                market = markets[FIRST]
                outcomes = market.get('outcomes', [])
                for outcome in outcomes:
                    bookmaker_name = bookmaker.get('title', '')
                    outcome_name = outcome.get('name', '')
                    outcome_price = outcome.get('price', 0)
                    
                    if not best_odds:
                        best_odds = [[bookmaker_name, outcome_name, outcome_price]]
                    else:
                        outcome_exists = False
                        for i, best_odd in enumerate(best_odds):
                            if outcome_name == best_odd[NAME_INDEX]:
                                outcome_exists = True
                                if outcome_price > best_odd[ODDS_INDEX]:
                                    best_odds[i] = [bookmaker_name, outcome_name, outcome_price]
                                break
                        if not outcome_exists:
                            best_odds.append([bookmaker_name, outcome_name, outcome_price])
        
        self.num_outcomes = len(best_odds)
        self.best_odds = best_odds
        return best_odds
   
    def arbitrage(self):
        total_arbitrage_percentage = 0
        for odds in self.best_odds:
            total_arbitrage_percentage += (1.0 / odds[ODDS_INDEX])
           
        self.total_arbitrage_percentage = total_arbitrage_percentage
       
        if total_arbitrage_percentage < 1:
            return True
        return False
   
    def convert_decimal_to_american(self):
        best_odds = self.best_odds
        for odds in best_odds:
            decimal = odds[ODDS_INDEX]
            if decimal >= 2:
                american = (decimal - 1) * 100
            elif decimal < 2:
                american = -100 / (decimal - 1)
            odds[ODDS_INDEX] = round(american, 2)
        return best_odds
     
    def calculate_arbitrage_bets(self, BET_SIZE):
        bet_amounts = []
        for outcome in range(self.num_outcomes):
            individual_arbitrage_percentage = 1 / self.best_odds[outcome][ODDS_INDEX]
            bet_amount = (BET_SIZE * individual_arbitrage_percentage) / self.total_arbitrage_percentage
            bet_amounts.append(round(bet_amount, 2))
       
        self.bet_amounts = bet_amounts
        self.expected_earnings = (BET_SIZE / self.total_arbitrage_percentage) - BET_SIZE
        return bet_amounts

def calculate_events(odds_response,BET_SIZE):
    events = []
    for data in odds_response:
        event = Event(data)
        event.find_best_odds()
        if event.arbitrage():
            event.calculate_arbitrage_bets(BET_SIZE)
            event.convert_decimal_to_american()
            events.append(event)
    return events

def create_arbitrage_df(arbitrage_events):
    MAX_OUTCOMES = max([event.num_outcomes for event in arbitrage_events])
    ARBITRAGE_EVENTS_COUNT = len(arbitrage_events)
    my_columns = ['ID', 'Sport Key', 'Expected Earnings'] + list(np.array([[f'Bookmaker #{outcome}', f'Name #{outcome}', f'Odds #{outcome}', f'Amount to Buy #{outcome}'] for outcome in range(1, MAX_OUTCOMES + 1)]).flatten())
    dataframe = pd.DataFrame(columns=my_columns)
    return dataframe, MAX_OUTCOMES, ARBITRAGE_EVENTS_COUNT

def write_data_to_excel(arbitrage_events, dataframe):
    for event in arbitrage_events:
        row = []
        row.append(event.id)
        row.append(event.sport_key)
        row.append(round(event.expected_earnings, 2))
        for index, outcome in enumerate(event.best_odds):
            row.append(outcome[BOOKMAKER_INDEX])
            row.append(outcome[NAME_INDEX])
            row.append(outcome[ODDS_INDEX])
            row.append(event.bet_amounts[index])
        while len(row) < len(dataframe.columns):
            row.append('N/A')
        dataframe.loc[len(dataframe.index)] = row

    writer = pd.ExcelWriter('bets.xlsx')
    dataframe.to_excel(writer, index=False)
    writer.close()

# %% [markdown]
# ### Formatting the Excel File

# %%
def format_excel(MAX_OUTCOMES, ARBITRAGE_EVENTS_COUNT):
    BLACK = '000000'
    LIGHT_GREY = 'D6D6D6'
    DARK_GREY = '9F9F9F'
    RED = 'FEA0A0'
    BLUE = 'A0CEFE'
    YELLOW = 'FFE540'

    COLORS = [RED, BLUE]

    ID_COLUMN_FILL = PatternFill(fill_type='solid', start_color=DARK_GREY, end_color=DARK_GREY)
    SPORT_KEY_COLUMN_FILL = PatternFill(fill_type='solid', start_color=LIGHT_GREY, end_color=LIGHT_GREY)
    EXPECTED_EARNINGS_COLUMN_FILL = PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)

    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='bottom', indent=0)

    TOP_ROW_BORDER = Border(bottom=Side(border_style='thick', color=BLACK))
    NORMAL_ROW_BORDER = Border(top=Side(border_style='thin', color=LIGHT_GREY), bottom=Side(border_style='thin', color=DARK_GREY))

    wb = load_workbook('bets.xlsx')
    ws = wb.active
    ws.title = 'Upcoming'
    # changing width
    for col in range(1, 26):
        ws.column_dimensions[chr(col + 64)].width = 20

    for cell in ws['A']:
        cell.fill = ID_COLUMN_FILL
        cell.alignment = CENTER_ALIGNMENT
        
    for cell in ws['B']:
        cell.fill = SPORT_KEY_COLUMN_FILL
        cell.alignment = CENTER_ALIGNMENT
        
    for cell in ws['C']:
        cell.fill = EXPECTED_EARNINGS_COLUMN_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.number_format = numbers.BUILTIN_FORMATS[7]

    START_INDEX = 'D'
    for index in range(MAX_OUTCOMES):
        for col in ws[START_INDEX : chr(ord(START_INDEX) + 3)]:
            for cell in col:
                color = COLORS[int(index % 2)]
                cell.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
                cell.alignment = CENTER_ALIGNMENT
                if cell.column % 4 == 3:
                    cell.number_format = numbers.BUILTIN_FORMATS[7]
                
        START_INDEX = chr(ord(START_INDEX) + 4)

    for cell in ws['1']:
        cell.border = TOP_ROW_BORDER

    for row in range(2, ARBITRAGE_EVENTS_COUNT + 2):
        for cell in ws[str(row)]:
            cell.border = NORMAL_ROW_BORDER
        
    wb.save('upcoming_events_bets.xlsx')

# %%



